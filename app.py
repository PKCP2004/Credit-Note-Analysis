
import io
import re
import zipfile
from datetime import datetime
from difflib import SequenceMatcher

import pandas as pd
import streamlit as st

st.set_page_config(
    page_title="Credit Note Mapping Tool",
    page_icon="🔎",
    layout="wide",
)

# -----------------------------
# Helpers
# -----------------------------
def clean_text(x):
    if pd.isna(x):
        return ""
    return re.sub(r"\s+", " ", str(x).strip())

def norm_doc(x):
    s = clean_text(x).upper()
    return re.sub(r"[^A-Z0-9]", "", s)

def to_num(x):
    if pd.isna(x) or x == "":
        return 0.0
    try:
        return float(str(x).replace(",", "").replace("₹", "").strip())
    except Exception:
        return 0.0

def parse_date(x):
    if pd.isna(x) or x == "":
        return pd.NaT
    return pd.to_datetime(x, dayfirst=True, errors="coerce")

def find_header_row(raw, required_terms):
    """Find the row containing the main column labels."""
    for i in range(min(len(raw), 20)):
        vals = [clean_text(v).lower() for v in raw.iloc[i].tolist()]
        joined = " | ".join(vals)
        if all(term.lower() in joined for term in required_terms):
            return i
    return None

def make_unique(cols):
    seen = {}
    out = []
    for c in cols:
        c = clean_text(c) or "Unnamed"
        seen[c] = seen.get(c, 0) + 1
        out.append(c if seen[c] == 1 else f"{c}_{seen[c]}")
    return out

def read_sheet(raw, sheet_name):
    # Special handling for GSTR-2B's two-row headers.
    if sheet_name == "B2B":
        header = find_header_row(raw, ["GSTIN of supplier", "Invoice number", "Taxable Value"])
        if header is None:
            return pd.DataFrame()
        h1 = raw.iloc[header].tolist()
        h2 = raw.iloc[header + 1].tolist() if header + 1 < len(raw) else [None] * raw.shape[1]
        cols = []
        for a, b in zip(h1, h2):
            a, b = clean_text(a), clean_text(b)
            cols.append(b if not a and b else (f"{a} - {b}" if a and b and a != b else a or b))
        df = raw.iloc[header + 2:].copy()
        df.columns = make_unique(cols)
        return df.dropna(how="all")

    if sheet_name == "B2B-CDNR":
        header = find_header_row(raw, ["GSTIN of supplier", "Note number", "Taxable Value"])
        if header is None:
            return pd.DataFrame()
        h1 = raw.iloc[header].tolist()
        h2 = raw.iloc[header + 1].tolist() if header + 1 < len(raw) else [None] * raw.shape[1]
        cols = []
        for a, b in zip(h1, h2):
            a, b = clean_text(a), clean_text(b)
            cols.append(b if not a and b else (f"{a} - {b}" if a and b and a != b else a or b))
        df = raw.iloc[header + 2:].copy()
        df.columns = make_unique(cols)
        return df.dropna(how="all")

    return pd.DataFrame()

def load_workbook_bytes(file_bytes, filename):
    """Return {sheet: raw dataframe}. Supports xlsx/xlsm and ZIP of workbooks."""
    workbooks = []
    if filename.lower().endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            for name in z.namelist():
                if name.lower().endswith((".xlsx", ".xlsm", ".xls")) and not name.startswith("__MACOSX"):
                    workbooks.append((name, z.read(name)))
    else:
        workbooks.append((filename, file_bytes))

    all_b2b = []
    all_cn = []
    sources = []

    for wb_name, wb_bytes in workbooks:
        try:
            xls = pd.ExcelFile(io.BytesIO(wb_bytes))
            sheets = xls.sheet_names
        except Exception as e:
            sources.append((wb_name, f"ERROR: {e}"))
            continue

        found = []
        for s in sheets:
            if s not in ("B2B", "B2B-CDNR"):
                continue
            try:
                raw = pd.read_excel(io.BytesIO(wb_bytes), sheet_name=s, header=None)
                parsed = read_sheet(raw, s)
                if parsed.empty:
                    continue
                parsed["_source_file"] = wb_name
                parsed["_source_sheet"] = s
                if s == "B2B":
                    all_b2b.append(parsed)
                    found.append("B2B")
                else:
                    all_cn.append(parsed)
                    found.append("B2B-CDNR")
            except Exception as e:
                sources.append((f"{wb_name}:{s}", f"ERROR: {e}"))
        sources.append((wb_name, ", ".join(found) if found else "No B2B/CDNR sheet"))

    b2b = pd.concat(all_b2b, ignore_index=True) if all_b2b else pd.DataFrame()
    cn = pd.concat(all_cn, ignore_index=True) if all_cn else pd.DataFrame()
    return b2b, cn, sources

def pick_col(df, *names):
    lookup = {clean_text(c).lower(): c for c in df.columns}
    for n in names:
        if n.lower() in lookup:
            return lookup[n.lower()]
    for c in df.columns:
        lc = clean_text(c).lower()
        if any(n.lower() in lc for n in names):
            return c
    return None

def standardize_b2b(df):
    if df.empty:
        return pd.DataFrame()
    mapping = {
        "month": pick_col(df, "GSTR-2B Month"),
        "gstin": pick_col(df, "GSTIN of supplier"),
        "supplier": pick_col(df, "Trade/Legal name"),
        "inv_no": pick_col(df, "Invoice number"),
        "inv_type": pick_col(df, "Invoice type"),
        "inv_date": pick_col(df, "Invoice Date"),
        "inv_value": pick_col(df, "Invoice Value(₹)", "Invoice Value"),
        "taxable": pick_col(df, "Taxable Value (₹)"),
        "igst": pick_col(df, "Integrated Tax(₹)"),
        "cgst": pick_col(df, "Central Tax(₹)"),
        "sgst": pick_col(df, "State/UT Tax(₹)"),
        "cess": pick_col(df, "Cess(₹)"),
        "itc": pick_col(df, "ITC Availability"),
    }
    missing = [k for k, v in mapping.items() if v is None and k in ("gstin","inv_no","taxable")]
    if missing:
        return pd.DataFrame()

    out = pd.DataFrame()
    for k, c in mapping.items():
        if c is not None:
            out[k] = df[c]
        else:
            out[k] = ""
    for c in ["gstin","supplier","inv_no","inv_type","month","itc"]:
        out[c] = out[c].map(clean_text)
    for c in ["inv_value","taxable","igst","cgst","sgst","cess"]:
        out[c] = out[c].map(to_num)
    out["inv_date"] = out["inv_date"].map(parse_date)
    out["doc_norm"] = out["inv_no"].map(norm_doc)
    out["tax_total"] = out[["igst","cgst","sgst","cess"]].sum(axis=1)
    out["source_file"] = df["_source_file"].values
    return out.dropna(subset=["gstin","inv_no"]).reset_index(drop=True)

def standardize_cn(df):
    if df.empty:
        return pd.DataFrame()
    mapping = {
        "month": pick_col(df, "GSTR-2B Month"),
        "gstin": pick_col(df, "GSTIN of supplier"),
        "supplier": pick_col(df, "Trade/Legal name"),
        "cn_no": pick_col(df, "Note number"),
        "note_type": pick_col(df, "Note type"),
        "note_date": pick_col(df, "Note date"),
        "cn_value": pick_col(df, "Note Value (₹)"),
        "taxable": pick_col(df, "Taxable Value (₹)"),
        "igst": pick_col(df, "Integrated Tax(₹)"),
        "cgst": pick_col(df, "Central Tax(₹)"),
        "sgst": pick_col(df, "State/UT Tax(₹)"),
        "cess": pick_col(df, "Cess(₹)"),
        "itc_avail": pick_col(df, "ITC Availability"),
        "itc_reduce": pick_col(df, "Whether ITC to be reduced (Taxpayer's Input)"),
        "taxpayer_igst": pick_col(df, "Amount declared by taxpayer for ITC reduction - Integrated Tax(₹)"),
        "taxpayer_cgst": pick_col(df, "Amount declared by taxpayer for ITC reduction - Central Tax(₹)"),
        "taxpayer_sgst": pick_col(df, "Amount declared by taxpayer for ITC reduction - State/UT Tax(₹)"),
        "taxpayer_cess": pick_col(df, "Amount declared by taxpayer for ITC reduction - Cess(₹)"),
    }
    if any(mapping[k] is None for k in ("gstin","cn_no","note_type","taxable")):
        return pd.DataFrame()

    out = pd.DataFrame()
    for k, c in mapping.items():
        out[k] = df[c] if c is not None else ""
    for c in ["gstin","supplier","cn_no","note_type","month","itc_avail","itc_reduce"]:
        out[c] = out[c].map(clean_text)
    for c in ["cn_value","taxable","igst","cgst","sgst","cess",
              "taxpayer_igst","taxpayer_cgst","taxpayer_sgst","taxpayer_cess"]:
        out[c] = out[c].map(to_num)
    out["note_date"] = out["note_date"].map(parse_date)
    out["doc_norm"] = out["cn_no"].map(norm_doc)
    out["tax_total"] = out[["igst","cgst","sgst","cess"]].sum(axis=1)
    out["source_file"] = df["_source_file"].values
    # Only original credit notes from the original CDN/CDNR sheet.
    out = out[out["note_type"].str.contains("credit", case=False, na=False)].copy()
    return out.reset_index(drop=True)

def similarity(a, b):
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()

def candidate_score(cn, inv):
    score = 0.0
    reasons = []

    if clean_text(cn["gstin"]).upper() == clean_text(inv["gstin"]).upper():
        score += 35
        reasons.append("Supplier GSTIN exact")

    # Exact taxable amount is a strong signal, but is not proof by itself.
    tax_diff = abs(cn["taxable"] - inv["taxable"])
    if tax_diff <= 0.01:
        score += 30
        reasons.append("Taxable value exact")
    elif max(abs(cn["taxable"]), abs(inv["taxable"]), 1) > 0:
        pct = tax_diff / max(abs(cn["taxable"]), abs(inv["taxable"]), 1)
        if pct <= 0.01:
            score += 22
            reasons.append("Taxable value within 1%")
        elif pct <= 0.05:
            score += 10
            reasons.append("Taxable value within 5%")

    tax_diff_total = abs(cn["tax_total"] - inv["tax_total"])
    if tax_diff_total <= 0.01:
        score += 20
        reasons.append("Tax amount exact")
    elif tax_diff_total <= max(1, abs(cn["tax_total"]) * 0.01):
        score += 12
        reasons.append("Tax amount within 1%")

    # Tax component pattern
    components = [("IGST", cn["igst"], inv["igst"]),
                  ("CGST", cn["cgst"], inv["cgst"]),
                  ("SGST", cn["sgst"], inv["sgst"])]
    component_matches = sum(abs(a-b) <= 0.01 for _,a,b in components)
    if component_matches == 3:
        score += 10
        reasons.append("IGST/CGST/SGST pattern exact")

    # Date proximity is weak evidence because a CN can be issued months later.
    if pd.notna(cn["note_date"]) and pd.notna(inv["inv_date"]):
        days = abs((cn["note_date"] - inv["inv_date"]).days)
        if days <= 30:
            score += 5
            reasons.append("Invoice/CN within 30 days")
        elif days <= 90:
            score += 2
            reasons.append("Invoice/CN within 90 days")

    return min(score, 100.0), reasons

def map_credit_notes(cns, invoices, max_candidates=5, tolerance=0.01):
    results = []
    if cns.empty:
        return pd.DataFrame()

    # Pre-index invoices by GSTIN for speed and safer matching.
    groups = {k: g for k, g in invoices.groupby(invoices["gstin"].str.upper(), sort=False)}

    for _, cn in cns.iterrows():
        pool = groups.get(clean_text(cn["gstin"]).upper(), pd.DataFrame()).copy()

        if pool.empty:
            results.append(make_result(cn, None, "UNMAPPED", 0, "No invoice found for same supplier GSTIN", []))
            continue

        scored = []
        for _, inv in pool.iterrows():
            s, reasons = candidate_score(cn, inv)
            scored.append((s, reasons, inv))
        scored.sort(key=lambda x: x[0], reverse=True)

        top = scored[:max_candidates]
        best_score, best_reasons, best_inv = top[0]

        # "Exact" requires amount/tax alignment, not just a high generic score.
        exact = (
            abs(cn["taxable"] - best_inv["taxable"]) <= tolerance
            and abs(cn["tax_total"] - best_inv["tax_total"]) <= tolerance
        )

        # If several invoices have essentially the same score, force review.
        close = [x for x in top if x[0] >= best_score - 3 and x[0] >= 70]
        if exact and best_score >= 80 and len(close) == 1:
            status = "MAPPED"
            match_type = "Exact amount/tax + unique candidate"
        elif best_score >= 75:
            status = "REVIEW"
            match_type = "Strong candidate - manual confirmation"
        elif best_score >= 55:
            status = "REVIEW"
            match_type = "Possible candidate - manual confirmation"
        else:
            status = "UNMAPPED"
            match_type = "No sufficiently strong candidate"

        results.append(make_result(cn, best_inv if status != "UNMAPPED" else None,
                                   status, round(best_score,1), match_type, best_reasons,
                                   alternatives=top[1:]))

    return pd.DataFrame(results)

def make_result(cn, inv, status, confidence, match_type, reasons, alternatives=None):
    alternatives = alternatives or []
    alt_text = " | ".join(
        f"{x[2]['inv_no']} ({x[0]:.0f}%)" for x in alternatives[:4]
    )
    if inv is None:
        inv_no = ""
        inv_date = ""
        inv_taxable = ""
        inv_tax = ""
        inv_file = ""
    else:
        inv_no = inv["inv_no"]
        inv_date = inv["inv_date"]
        inv_taxable = inv["taxable"]
        inv_tax = inv["tax_total"]
        inv_file = inv["source_file"]

    cn_itc = cn["tax_total"]
    return {
        "GSTR-2B Month": cn["month"],
        "Supplier GSTIN": cn["gstin"],
        "Supplier": cn["supplier"],
        "Credit Note No": cn["cn_no"],
        "CN Date": cn["note_date"],
        "CN Value": cn["cn_value"],
        "CN Taxable": cn["taxable"],
        "CN IGST": cn["igst"],
        "CN CGST": cn["cgst"],
        "CN SGST": cn["sgst"],
        "CN Cess": cn["cess"],
        "Original Invoice Candidate": inv_no,
        "Original Invoice Date": inv_date,
        "Original Taxable": inv_taxable,
        "Original GST": inv_tax,
        "Match Type": match_type,
        "Confidence %": confidence,
        "Status": status,
        "CN ITC Impact": cn_itc,
        "Matching Reasons": "; ".join(reasons),
        "Other Candidates": alt_text,
        "Source File": cn["source_file"],
        "Invoice Source File": inv_file,
    }


def build_upload_template():
    """Generate the standardized B2B + B2B-CDNR workbook used by the tool."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    b2b_headers=[
        "GSTR-2B Month","GSTIN of supplier","Trade/Legal name","Invoice number",
        "Invoice type","Invoice Date","Invoice Value(₹)","Place of Supply",
        "Reverse Charge","Taxable Value (₹)","Integrated Tax(₹)","Central Tax(₹)",
        "State/UT Tax(₹)","Cess(₹)","GSTR-1/1A/IFF/GSTR-5 Period",
        "GSTR-1/1A/IFF/GSTR-5 Filing Date","ITC Availability","Reason",
        "Applicable % of Tax Rate","Source","IRN","IRN Date"
    ]
    cdnr_headers=[
        "GSTR-2B Month","GSTIN of supplier","Trade/Legal name","Note number",
        "Note type","Note Supply Type","Note Date","Note Value (₹)",
        "Place of Supply","Reverse Charge","Taxable Value (₹)",
        "Integrated Tax(₹)","Central Tax(₹)","State/UT Tax(₹)","Cess(₹)",
        "Whether ITC to be reduced",
        "Amount declared by taxpayer for ITC reduction - Integrated Tax(₹)",
        "Amount declared by taxpayer for ITC reduction - Central Tax(₹)",
        "Amount declared by taxpayer for ITC reduction - State/UT Tax(₹)",
        "Amount declared by taxpayer for ITC reduction - Cess(₹)","Remarks",
        "GSTR-1/1A/IFF/GSTR-5 Period","GSTR-1/1A/IFF/GSTR-5 Filing Date",
        "ITC Availability","Reason","Applicable % of Tax Rate","Source","IRN","IRN Date"
    ]

    wb=Workbook()
    b2b=wb.active
    b2b.title="B2B"
    cdnr=wb.create_sheet("B2B-CDNR")
    ins=wb.create_sheet("Instructions")

    fill=PatternFill("solid",fgColor="8B0033")
    for ws,headers in [(b2b,b2b_headers),(cdnr,cdnr_headers)]:
        ws.append(headers)
        ws.freeze_panes="A2"
        ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}501"
        ws.row_dimensions[1].height=42
        for c in ws[1]:
            c.font=Font(bold=True,color="FFFFFF")
            c.fill=fill
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        for _ in range(500):
            ws.append([""]*len(headers))
        for i,h in enumerate(headers,1):
            ws.column_dimensions[get_column_letter(i)].width=min(max(len(h)+2,14),34)

    ins.append(["PUSHPAK KUMAR — Credit Note Mapping Tool"])
    ins.append(["Standard Upload Format"])
    ins.append(["Sheet","Instruction"])
    ins.append(["B2B","Paste one invoice per row. Keep columns unchanged."])
    ins.append(["B2B-CDNR","Paste one note per row. Credit Notes will be analysed; Debit Notes are excluded."])
    ins.append(["Dates","Use DD/MM/YYYY where possible."])
    ins.append(["Amounts","Use numeric values without ₹ symbols."])
    ins.append(["Important","Do not rename the sheets or change the column order."])
    ins.merge_cells("A1:B1"); ins.merge_cells("A2:B2")
    for cell in [ins["A1"],ins["A2"]]:
        cell.font=Font(bold=True,color="FFFFFF",size=16 if cell==ins["A1"] else 11)
        cell.fill=fill
    for c in ins[3]:
        c.font=Font(bold=True,color="FFFFFF"); c.fill=fill
    ins.column_dimensions["A"].width=24
    ins.column_dimensions["B"].width=105
    for row in ins.iter_rows():
        for c in row:
            c.alignment=Alignment(wrap_text=True,vertical="top")

    out=io.BytesIO()
    wb.save(out); out.seek(0)
    return out.getvalue()


# -----------------------------
# Professional Excel Export
# -----------------------------
def build_professional_excel(invoices, cns, mapping, source_log, source_name):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    MAROON="8B0033"; DARK="5C0022"; GREEN="E2F0D9"; YELLOW="FFF2CC"; RED="FCE4D6"
    GREY="666666"; WHITE="FFFFFF"; BORDER="D9D9E3"

    m=mapping.copy(); inv=invoices.copy()

    mapped=m[m["Status"]=="MAPPED"].copy()
    if not mapped.empty:
        link=(mapped.groupby(["Supplier GSTIN","Original Invoice Candidate"],dropna=False)
              .agg(Mapped_CN_Count=("Credit Note No","count"),
                   Mapped_CN_Numbers=("Credit Note No",lambda x:", ".join(map(str,x))),
                   Mapped_CN_Taxable=("CN Taxable","sum"),
                   Mapped_CN_ITC=("CN ITC Impact","sum"),
                   Lowest_Confidence=("Confidence %","min")).reset_index())
    else:
        link=pd.DataFrame(columns=["Supplier GSTIN","Original Invoice Candidate","Mapped_CN_Count",
                                   "Mapped_CN_Numbers","Mapped_CN_Taxable","Mapped_CN_ITC","Lowest_Confidence"])

    b2b=inv.merge(link,left_on=["gstin","inv_no"],
                  right_on=["Supplier GSTIN","Original Invoice Candidate"],how="left")
    for c in ["Mapped_CN_Count","Mapped_CN_Taxable","Mapped_CN_ITC","Lowest_Confidence"]:
        b2b[c]=b2b[c].fillna(0)
    b2b["Mapped_CN_Count"]=b2b["Mapped_CN_Count"].astype(int)
    b2b["Mapped_CN_Numbers"]=b2b["Mapped_CN_Numbers"].fillna("")
    b2b["Remaining ITC After CN"]=(b2b["tax_total"]-b2b["Mapped_CN_ITC"]).clip(lower=0)
    b2b["Excess CN ITC"]=(b2b["Mapped_CN_ITC"]-b2b["tax_total"]).clip(lower=0)
    b2b["CN Mapping Status"]=b2b["Mapped_CN_Count"].map(lambda x:"CREDIT NOTE MAPPED" if x else "NO MAPPED CREDIT NOTE")

    b2b_out=pd.DataFrame({
        "GSTR-2B Month":b2b["month"],"Supplier GSTIN":b2b["gstin"],"Supplier":b2b["supplier"],
        "Invoice No":b2b["inv_no"],"Invoice Type":b2b["inv_type"],"Invoice Date":b2b["inv_date"],
        "Invoice Value":b2b["inv_value"],"Taxable Value":b2b["taxable"],"IGST":b2b["igst"],
        "CGST":b2b["cgst"],"SGST":b2b["sgst"],"Cess":b2b["cess"],"Total GST / ITC":b2b["tax_total"],
        "ITC Availability":b2b["itc"],"Mapped CN Count":b2b["Mapped_CN_Count"],
        "Mapped Credit Note Nos":b2b["Mapped_CN_Numbers"],"Mapped CN Taxable":b2b["Mapped_CN_Taxable"],
        "Mapped CN ITC Impact":b2b["Mapped_CN_ITC"],"Remaining ITC After CN":b2b["Remaining ITC After CN"],
        "Excess CN ITC":b2b["Excess CN ITC"],"Lowest Mapping Confidence %":b2b["Lowest_Confidence"],
        "CN Mapping Status":b2b["CN Mapping Status"],"Source File":b2b["source_file"]
    })

    preferred=["GSTR-2B Month","Supplier GSTIN","Supplier","Credit Note No","CN Date","CN Value","CN Taxable",
               "CN IGST","CN CGST","CN SGST","CN Cess","Original Invoice Candidate","Original Invoice Date",
               "Original Taxable","Original GST","Match Type","Confidence %","Status","CN ITC Impact",
               "Matching Reasons","Other Candidates","Source File","Invoice Source File"]
    cn_out=m[[x for x in preferred if x in m.columns]].copy()

    if not mapped.empty:
        knock=(mapped.groupby(["Supplier GSTIN","Supplier","Original Invoice Candidate"],dropna=False)
               .agg(Credit_Note_Count=("Credit Note No","count"),
                    Credit_Note_Numbers=("Credit Note No",lambda x:", ".join(map(str,x))),
                    CN_Taxable=("CN Taxable","sum"),CN_ITC_Impact=("CN ITC Impact","sum"),
                    Original_Invoice_ITC=("Original GST","first"),
                    Lowest_Confidence=("Confidence %","min")).reset_index())
        knock["Remaining ITC"]=(knock["Original_Invoice_ITC"]-knock["CN_ITC_Impact"]).clip(lower=0)
        knock["Excess CN ITC"]=(knock["CN_ITC_Impact"]-knock["Original_Invoice_ITC"]).clip(lower=0)
    else:
        knock=pd.DataFrame(columns=["Supplier GSTIN","Supplier","Original Invoice Candidate","Credit_Note_Count",
                                    "Credit_Note_Numbers","CN_Taxable","CN_ITC_Impact","Original_Invoice_ITC",
                                    "Remaining ITC","Excess CN ITC","Lowest_Confidence"])
    review=m[m["Status"]!="MAPPED"].copy()

    mapped_n=int((m["Status"]=="MAPPED").sum()); review_n=int((m["Status"]=="REVIEW").sum())
    unmapped_n=int((m["Status"]=="UNMAPPED").sum()); total_cn=len(m)

    dashboard=pd.DataFrame({
        "Metric":["Source File","B2B Invoices Analysed","Credit Notes Analysed","High-Confidence Mapped",
                  "Manual Review","Unmapped","Mapping Coverage","Total Invoice ITC",
                  "Total Credit Note ITC Impact","Remaining ITC After Mapped CN","Excess CN ITC"],
        "Value":[source_name,len(b2b_out),total_cn,mapped_n,review_n,unmapped_n,
                  mapped_n/total_cn if total_cn else 0,b2b_out["Total GST / ITC"].sum(),
                  m["CN ITC Impact"].sum(),b2b_out["Remaining ITC After CN"].sum(),
                  b2b_out["Excess CN ITC"].sum()]
    })

    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as writer:
        dashboard.to_excel(writer,sheet_name="Dashboard",index=False,startrow=4)
        b2b_out.to_excel(writer,sheet_name="B2B Invoice Register",index=False)
        cn_out.to_excel(writer,sheet_name="Credit Note Mapping",index=False)
        knock.to_excel(writer,sheet_name="Knock-off",index=False)
        review.to_excel(writer,sheet_name="Review Queue",index=False)
        pd.DataFrame(source_log,columns=["File","Detected sheets / status"]).to_excel(writer,sheet_name="Source Info",index=False)
    buf.seek(0); wb=load_workbook(buf)

    def style(ws,title):
        ws.sheet_view.showGridLines=False
        ws.insert_rows(1,3)
        ws["A1"]=title; ws["A1"].font=Font(size=19,bold=True,color=WHITE)
        ws["A1"].fill=PatternFill("solid",fgColor=MAROON)
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=max(2,ws.max_column))
        ws["A2"]="Generated report from Credit Note Mapping Tool  •  pushpakkumar.com"
        ws["A2"].font=Font(size=10,italic=True,color=GREY)
        for c in ws[4]:
            c.font=Font(bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=DARK)
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.freeze_panes="A5"; ws.auto_filter.ref=ws.dimensions; ws.row_dimensions[4].height=34
        for col in ws.columns:
            letter=get_column_letter(col[0].column)
            vals=[str(x.value or "") for x in col[:100]]
            ws.column_dimensions[letter].width=min(max(max(map(len,vals),default=10)+2,12),40)
        for row in ws.iter_rows(min_row=5):
            for c in row: c.alignment=Alignment(vertical="top",wrap_text=True)
        ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1
        ws.sheet_properties.pageSetUpPr.fitToPage=True
        ws.oddFooter.center.text="Generated report from Credit Note Mapping Tool  •  pushpakkumar.com"
        ws.oddFooter.right.text="Developed & Designed by Pushpak Kumar"

    d=wb["Dashboard"]; d.sheet_view.showGridLines=False
    d["A1"]="PUSHPAK KUMAR"; d["A1"].font=Font(size=22,bold=True,color=WHITE)
    d["A1"].fill=PatternFill("solid",fgColor=MAROON); d.merge_cells("A1:F1")
    d["A2"]="Credit Note Mapping & Knock-off  |  Generated from pushpakkumar.com"; d["A2"].font=Font(size=12,bold=True,color=DARK); d.merge_cells("A2:F2")
    d["A3"]="Professional working paper generated from the uploaded GSTR-2B workbook"; d["A3"].font=Font(italic=True,color=GREY); d.merge_cells("A3:F3")
    for c in d[5]:
        c.font=Font(bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=DARK)
    d["B11"].number_format="0.0%"
    for r in range(12,16): d[f"B{r}"].number_format='₹#,##0.00'
    d["D5"]="Status"; d["E5"]="Count"
    for c in d[5][3:5]: c.font=Font(bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=MAROON)
    for r,(s,n) in enumerate([("MAPPED",mapped_n),("REVIEW",review_n),("UNMAPPED",unmapped_n)],6):
        d[f"D{r}"]=s; d[f"E{r}"]=n
    chart=BarChart(); chart.title="Credit Note Mapping Status"; chart.y_axis.title="Credit Notes"
    chart.add_data(Reference(d,min_col=5,min_row=5,max_row=8),titles_from_data=True)
    chart.set_categories(Reference(d,min_col=4,min_row=6,max_row=8)); chart.height=7; chart.width=12; d.add_chart(chart,"D10")
    d.column_dimensions["A"].width=44; d.column_dimensions["B"].width=28
    d.oddFooter.center.text="Generated report from Credit Note Mapping Tool  •  pushpakkumar.com"; d.oddFooter.right.text="Developed & Designed by Pushpak Kumar"

    for name,title in [
        ("B2B Invoice Register","B2B Invoice Register — Credit Note Mapping"),
        ("Credit Note Mapping","Credit Note Mapping — Detailed Analysis"),
        ("Knock-off","Invoice-level Credit Note Knock-off"),
        ("Review Queue","Credit Notes Requiring Review"),
        ("Source Info","Source Information"),
    ]: style(wb[name],title)

    for name in ["B2B Invoice Register","Credit Note Mapping","Knock-off"]:
        ws=wb[name]; headers={c.value:c.column for c in ws[4]}
        for h,col in headers.items():
            if any(k in str(h) for k in ["Value","Taxable","IGST","CGST","SGST","Cess","GST","ITC","Impact","Remaining","Excess"]):
                for r in range(5,ws.max_row+1): ws.cell(r,col).number_format='₹#,##0.00'
            if "%" in str(h) or "Confidence" in str(h):
                for r in range(5,ws.max_row+1): ws.cell(r,col).number_format='0.0'

    # Direct status fills (no conditional-formatting XML required).
    status_fills = {"MAPPED": GREEN, "REVIEW": YELLOW, "UNMAPPED": RED}
    ws = wb["Credit Note Mapping"]
    headers = {c.value: c.column for c in ws[4]}
    if "Status" in headers:
        col = headers["Status"]
        for r in range(5, ws.max_row + 1):
            value = str(ws.cell(r, col).value or "")
            if value in status_fills:
                ws.cell(r, col).fill = PatternFill("solid", fgColor=status_fills[value])
    ws = wb["B2B Invoice Register"]
    headers = {c.value: c.column for c in ws[4]}
    if "CN Mapping Status" in headers:
        col = headers["CN Mapping Status"]
        for r in range(5, ws.max_row + 1):
            value = str(ws.cell(r, col).value or "")
            if value == "CREDIT NOTE MAPPED":
                ws.cell(r, col).fill = PatternFill("solid", fgColor=GREEN)

    wb.properties.creator="Pushpak Kumar"
    wb.properties.title="Credit Note Mapping & Knock-off Report"
    out=io.BytesIO(); wb.save(out); out.seek(0); return out.getvalue()



# -----------------------------
# Fresh Professional UI
# -----------------------------
st.set_page_config(
    page_title="CN Map | Pushpak Kumar",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
/* ---------- Base ---------- */
.stApp {
    background:
        radial-gradient(circle at 90% 0%, rgba(139,0,51,.10), transparent 28%),
        radial-gradient(circle at 0% 30%, rgba(44,62,80,.08), transparent 25%),
        #0b0d11;
    color:#eef0f4;
}
.block-container {
    max-width: 1500px;
    padding: 30px 42px 48px;
}
header[data-testid="stHeader"] { background:transparent; }
footer { visibility:hidden; }

/* ---------- Top nav ---------- */
.topbar {
    display:flex;
    align-items:center;
    justify-content:space-between;
    padding: 4px 0 30px;
}
.brand {
    display:flex;
    align-items:center;
    gap:12px;
}
.logo {
    width:38px;height:38px;border-radius:11px;
    display:flex;align-items:center;justify-content:center;
    background:linear-gradient(145deg,#9e174b,#69002b);
    border:1px solid rgba(255,255,255,.14);
    font-weight:900;font-size:13px;
    box-shadow:0 8px 22px rgba(139,0,51,.25);
}
.brand-title {font-weight:800;font-size:15px;letter-spacing:.3px;}
.brand-sub {font-size:10px;color:#8f96a3;letter-spacing:1.2px;margin-top:2px;}
.status-pill {
    border:1px solid #242a35;background:#11151c;color:#aeb6c4;
    border-radius:999px;padding:7px 12px;font-size:11px;
}

/* ---------- Hero ---------- */
.hero {
    position:relative;overflow:hidden;
    border:1px solid #242a34;
    border-radius:24px;
    padding:42px 44px 40px;
    background:
      linear-gradient(135deg,rgba(255,255,255,.035),rgba(255,255,255,.008)),
      #11141a;
    box-shadow:0 25px 70px rgba(0,0,0,.24);
}
.hero:after {
    content:"";
    position:absolute;right:-120px;top:-180px;
    width:420px;height:420px;border-radius:50%;
    border:1px solid rgba(139,0,51,.22);
    box-shadow:0 0 0 55px rgba(139,0,51,.025),0 0 0 110px rgba(139,0,51,.018);
}
.eyebrow {
    color:#d34a79;font-size:11px;font-weight:800;
    letter-spacing:1.7px;text-transform:uppercase;margin-bottom:14px;
}
.hero h1 {
    font-size:43px;line-height:1.03;margin:0;
    letter-spacing:-1.7px;color:#fff;font-weight:850;
}
.hero p {
    max-width:760px;color:#969eac;font-size:14px;
    line-height:1.65;margin:16px 0 0;
}

/* ---------- Workflow ---------- */
.workflow {
    display:grid;grid-template-columns:repeat(4,1fr);
    gap:10px;margin:16px 0 22px;
}
.flow {
    border:1px solid #222833;background:#10141a;
    border-radius:14px;padding:13px 15px;
    display:flex;align-items:center;gap:11px;
}
.flow-no {
    width:27px;height:27px;border-radius:8px;
    display:flex;align-items:center;justify-content:center;
    background:#1c212a;color:#d34a79;font-weight:800;font-size:11px;
}
.flow b{font-size:12px;color:#e8ebef}.flow span{font-size:10px;color:#737c89;display:block;margin-top:2px}

/* ---------- Main cards ---------- */
.card {
    background:#10141a;
    border:1px solid #222833;
    border-radius:19px;
    padding:24px;
    box-shadow:0 15px 45px rgba(0,0,0,.15);
}
.card-head {
    display:flex;align-items:flex-start;justify-content:space-between;
    gap:20px;margin-bottom:20px;
}
.card-title {font-size:16px;font-weight:800;color:#f1f3f6;}
.card-desc {font-size:11px;color:#777f8c;margin-top:5px;line-height:1.5;}
.kicker {
    color:#737d8c;font-size:10px;font-weight:800;
    letter-spacing:1.3px;text-transform:uppercase;
}
.drop-zone {
    border:1px dashed #3a414d;border-radius:15px;
    padding:24px;background:#0c1015;
}
.helper {
    color:#737c89;font-size:10px;line-height:1.5;margin-top:9px;
}

/* ---------- Metrics ---------- */
.metric-wrap {
    border:1px solid #222833;background:#10141a;
    border-radius:15px;padding:15px 16px;
}
.metric-label{font-size:10px;color:#777f8c;text-transform:uppercase;letter-spacing:1px;}
.metric-value{font-size:22px;font-weight:800;color:#f6f7f9;margin-top:5px;}
.metric-accent{font-size:10px;color:#b53b68;margin-top:3px}

/* ---------- Buttons ---------- */
.stDownloadButton button, .stButton button {
    border-radius:10px !important;
    border:1px solid #343b47 !important;
    background:#171b22 !important;
    color:#f2f3f5 !important;
    font-weight:700 !important;
    min-height:42px !important;
}
.stDownloadButton button:hover,.stButton button:hover {
    border-color:#9b1b4e !important;
    background:#1c151a !important;
}
[data-testid="stFileUploaderDropzone"] {
    background:#0c1015 !important;border:0 !important;
}
[data-testid="stFileUploaderDropzone"] button {
    background:#8b0033 !important;border-color:#8b0033 !important;color:white !important;
}

/* ---------- Tabs / tables ---------- */
.stTabs [data-baseweb="tab-list"] {
    gap:4px;background:#0d1116;border:1px solid #222833;
    padding:5px;border-radius:12px;
}
.stTabs [data-baseweb="tab"] {
    color:#737c89;padding:9px 15px;border-radius:8px;font-size:11px;
}
.stTabs [aria-selected="true"] {
    background:#1a1519 !important;color:#f0dce4 !important;
}
div[data-testid="stDataFrame"] {
    border:1px solid #222833;border-radius:12px;overflow:hidden;
}
div[data-testid="stMetric"] {
    background:#10141a;border:1px solid #222833;
    border-radius:15px;padding:12px 15px;
}
div[data-testid="stMetricLabel"] {color:#7d8693 !important;font-size:10px !important;}
div[data-testid="stMetricValue"] {color:#f4f5f7 !important;font-size:21px !important;}

/* ---------- Empty state / footer ---------- */
.empty {
    text-align:center;padding:30px 10px;color:#69727f;
}
.empty strong{color:#dfe2e7}
.footer {
    text-align:center;color:#545c68;font-size:10px;
    letter-spacing:.6px;margin-top:35px;
}
.footer b{color:#9b1b4e}
@media(max-width:900px){
 .block-container{padding:20px 18px 40px}
 .workflow{grid-template-columns:1fr 1fr}
 .hero h1{font-size:32px}
}
</style>
""", unsafe_allow_html=True)

# Top navigation
st.markdown("""
<div class="topbar">
  <div class="brand">
    <div class="logo">PK</div>
    <div>
      <div class="brand-title">PUSHPAK KUMAR</div>
      <div class="brand-sub">GST AUTOMATION LAB</div>
    </div>
  </div>
  <div class="status-pill">● LOCAL PROCESSING · EXCEL</div>
</div>
""", unsafe_allow_html=True)

# Hero
st.markdown("""
<section class="hero">
  <div class="eyebrow">GST • ITC CONTROL • AUTOMATION</div>
  <h1>Credit Note<br>Mapping Engine</h1>
  <p>
    Turn your GSTR-2B workbook into an invoice-linked working paper.
    The engine identifies credit notes, evaluates original-invoice candidates,
    separates review cases, and prepares an Excel knock-off schedule.
  </p>
</section>
""", unsafe_allow_html=True)

st.markdown("""
<div class="workflow">
 <div class="flow"><div class="flow-no">01</div><div><b>Upload</b><span>Portal workbook</span></div></div>
 <div class="flow"><div class="flow-no">02</div><div><b>Analyse</b><span>Invoices & notes</span></div></div>
 <div class="flow"><div class="flow-no">03</div><div><b>Validate</b><span>Match confidence</span></div></div>
 <div class="flow"><div class="flow-no">04</div><div><b>Export</b><span>Working paper</span></div></div>
</div>
""", unsafe_allow_html=True)

# Upload workspace
left,right=st.columns([1.05,1.95],gap="large")

with left:
    st.markdown("""
    <div class="card">
      <div class="card-head">
        <div>
          <div class="kicker">START HERE</div>
          <div class="card-title" style="margin-top:5px">Standard upload format</div>
          <div class="card-desc">Use the same workbook structure every time for consistent analysis.</div>
        </div>
      </div>
    """,unsafe_allow_html=True)
    st.download_button(
        "Download Excel Format",
        data=build_upload_template(),
        file_name="Credit_Note_Mapping_Upload_Format.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    st.markdown("""
      <div class="helper">
      Required sheets: <b>B2B</b> and <b>B2B-CDNR</b>.<br>
      Credit Notes are analysed; Debit Notes are excluded from mapping.
      </div></div>
    """,unsafe_allow_html=True)

with right:
    st.markdown("""
    <div class="card">
      <div class="card-head">
        <div>
          <div class="kicker">INPUT WORKSPACE</div>
          <div class="card-title" style="margin-top:5px">Upload GSTR-2B data</div>
          <div class="card-desc">Drop your completed workbook here. ZIP uploads are also supported where applicable.</div>
        </div>
      </div>
      <div class="drop-zone">
    """,unsafe_allow_html=True)
    uploaded=st.file_uploader(
        "Upload workbook",
        type=["xlsx","xlsm","xls","zip"],
        label_visibility="collapsed"
    )
    st.markdown("</div>",unsafe_allow_html=True)

st.markdown("<div style='height:8px'></div>",unsafe_allow_html=True)

if not uploaded:
    st.markdown("""
    <div class="card">
      <div class="empty">
        <strong>Your analysis workspace is ready.</strong><br>
        Download the format, populate B2B + B2B-CDNR, then upload it above.
      </div>
    </div>
    <div class="footer">DESIGNED & DEVELOPED BY <b>PUSHPAK KUMAR</b> · GST AUTOMATION LAB</div>
    """,unsafe_allow_html=True)
    st.stop()

# Process
with st.spinner("Analysing workbook and building candidate matches..."):
    try:
        b2b_raw,cn_raw,source_log=load_workbook_bytes(uploaded.getvalue(),uploaded.name)
        invoices=standardize_b2b(b2b_raw)
        cns=standardize_cn(cn_raw)
    except Exception as exc:
        st.error("The workbook could not be processed.")
        st.exception(exc)
        st.stop()

if invoices.empty:
    st.error("No usable B2B invoice data was detected. Please download and use the standard upload format.")
    st.stop()
if cns.empty:
    st.warning("B2B was loaded, but no Credit Note rows were detected in B2B-CDNR.")
    st.stop()

try:
    mapping=map_credit_notes(cns,invoices)
except Exception as exc:
    st.error("The mapping engine encountered an unexpected issue.")
    st.exception(exc)
    st.stop()

counts=mapping["Status"].value_counts()
mapped_n=int(counts.get("MAPPED",0))
review_n=int(counts.get("REVIEW",0))
unmapped_n=int(counts.get("UNMAPPED",0))
coverage=mapped_n/len(mapping) if len(mapping) else 0

st.markdown("""
<div class="card" style="margin-top:10px;margin-bottom:18px">
 <div class="kicker">ANALYSIS COMPLETE</div>
 <div class="card-title" style="margin-top:5px">Working paper generated from your upload</div>
</div>
""",unsafe_allow_html=True)

# Metrics
metrics=[
    ("B2B INVOICES",f"{len(invoices):,}","records"),
    ("CREDIT NOTES",f"{len(cns):,}","analysed"),
    ("MAPPED",f"{mapped_n:,}","confirmed candidates"),
    ("REVIEW",f"{review_n+unmapped_n:,}","needs attention"),
    ("COVERAGE",f"{coverage:.1%}","mapping coverage"),
]
mcols=st.columns(5)
for col,(label,value,sub) in zip(mcols,metrics):
    with col:
        st.markdown(f'<div class="metric-wrap"><div class="metric-label">{label}</div><div class="metric-value">{value}</div><div class="metric-accent">{sub}</div></div>',unsafe_allow_html=True)

st.markdown("<div style='height:18px'></div>",unsafe_allow_html=True)

tab1,tab2,tab3,tab4=st.tabs([
    "Overview","B2B Invoice Register","Credit Note Mapping","Review Queue"
])

with tab1:
    a,b,c=st.columns(3)
    with a:
        st.markdown('<div class="metric-wrap"><div class="metric-label">INVOICE ITC</div><div class="metric-value">₹{:,.2f}</div><div class="metric-accent">Total B2B GST / ITC</div></div>'.format(invoices["tax_total"].sum()),unsafe_allow_html=True)
    with b:
        st.markdown('<div class="metric-wrap"><div class="metric-label">CN ITC IMPACT</div><div class="metric-value">₹{:,.2f}</div><div class="metric-accent">Mapped + review impact</div></div>'.format(mapping["CN ITC Impact"].sum()),unsafe_allow_html=True)
    with c:
        st.markdown('<div class="metric-wrap"><div class="metric-label">EXCEPTIONS</div><div class="metric-value">{:,}</div><div class="metric-accent">Review + unmapped</div></div>'.format(review_n+unmapped_n),unsafe_allow_html=True)

    st.markdown("<div style='height:18px'></div>",unsafe_allow_html=True)
    q1,q2=st.columns([1,1],gap="large")
    with q1:
        st.markdown('<div class="card"><div class="kicker">STATUS</div><div class="card-title" style="margin-top:5px">Mapping distribution</div>',unsafe_allow_html=True)
        st.bar_chart(mapping["Status"].value_counts())
        st.markdown("</div>",unsafe_allow_html=True)
    with q2:
        st.markdown('<div class="card"><div class="kicker">SUPPLIERS</div><div class="card-title" style="margin-top:5px">Credit note impact by supplier</div>',unsafe_allow_html=True)
        supplier_summary=(mapping.groupby(["Supplier GSTIN","Supplier"],dropna=False)
            .agg(Credit_Notes=("Credit Note No","count"),CN_Taxable=("CN Taxable","sum"),ITC_Impact=("CN ITC Impact","sum"))
            .reset_index().sort_values("ITC_Impact",ascending=False))
        st.dataframe(supplier_summary,use_container_width=True,hide_index=True)
        st.markdown("</div>",unsafe_allow_html=True)

with tab2:
    st.markdown('<div class="card"><div class="kicker">INVOICE CONTROL</div><div class="card-title" style="margin-top:5px">B2B Invoice Register</div><div class="card-desc">Every invoice remains visible. Mapped credit notes are aggregated against their original invoice.</div></div>',unsafe_allow_html=True)
    mapped=mapping[mapping["Status"]=="MAPPED"].copy()
    if not mapped.empty:
        link=(mapped.groupby(["Supplier GSTIN","Original Invoice Candidate"],dropna=False)
            .agg(Mapped_CN_Count=("Credit Note No","count"),Mapped_CN_Numbers=("Credit Note No",lambda x:", ".join(map(str,x))),
                 Mapped_CN_Taxable=("CN Taxable","sum"),Mapped_CN_ITC=("CN ITC Impact","sum")).reset_index())
        bv=invoices.rename(columns={"gstin":"Supplier GSTIN","inv_no":"Invoice No"}).merge(
            link,left_on=["Supplier GSTIN","Invoice No"],right_on=["Supplier GSTIN","Original Invoice Candidate"],how="left")
        for cc in ["Mapped_CN_Count","Mapped_CN_Taxable","Mapped_CN_ITC"]: bv[cc]=bv[cc].fillna(0)
        bv["Mapped_CN_Count"]=bv["Mapped_CN_Count"].astype(int); bv["Mapped_CN_Numbers"]=bv["Mapped_CN_Numbers"].fillna("")
        bv["Remaining ITC"]=(bv["tax_total"]-bv["Mapped_CN_ITC"]).clip(lower=0)
        bv=bv[["Supplier GSTIN","supplier","Invoice No","inv_date","taxable","tax_total","Mapped_CN_Count","Mapped_CN_Numbers","Mapped_CN_Taxable","Mapped_CN_ITC","Remaining ITC"]].rename(
            columns={"supplier":"Supplier","inv_date":"Invoice Date","taxable":"Invoice Taxable","tax_total":"Invoice GST / ITC",
                     "Mapped_CN_Count":"CN Count","Mapped_CN_Numbers":"Credit Note Nos","Mapped_CN_Taxable":"CN Taxable","Mapped_CN_ITC":"CN ITC Impact"})
    else:
        bv=invoices.copy()
    st.dataframe(bv,use_container_width=True,hide_index=True)

with tab3:
    st.markdown('<div class="card"><div class="kicker">MATCH ENGINE</div><div class="card-title" style="margin-top:5px">Credit Note → Original Invoice</div><div class="card-desc">Review confidence, match reasons and candidate invoice references.</div></div>',unsafe_allow_html=True)
    statuses=st.multiselect("Status filter",["MAPPED","REVIEW","UNMAPPED"],default=["MAPPED","REVIEW","UNMAPPED"])
    st.dataframe(mapping[mapping["Status"].isin(statuses)],use_container_width=True,hide_index=True)

with tab4:
    exceptions=mapping[mapping["Status"]!="MAPPED"].copy()
    st.markdown('<div class="card"><div class="kicker">EXCEPTIONS</div><div class="card-title" style="margin-top:5px">Human review queue</div><div class="card-desc">These items were deliberately not treated as confirmed knock-offs.</div></div>',unsafe_allow_html=True)
    if exceptions.empty:
        st.success("No exceptions. All credit notes have a confirmed mapping.")
    else:
        st.dataframe(exceptions,use_container_width=True,hide_index=True)

st.markdown("<div style='height:18px'></div>",unsafe_allow_html=True)
st.markdown("""
<div class="card">
  <div class="kicker">DELIVERABLE</div>
  <div class="card-title" style="margin-top:5px">Professional Excel working paper</div>
  <div class="card-desc">Branded dashboard, invoice register, detailed CN mapping, knock-off schedule, review queue and source information.</div>
</div>
""",unsafe_allow_html=True)

excel_bytes=build_professional_excel(invoices,cns,mapping,source_log,uploaded.name)
st.download_button(
    "Download Professional Excel Working Paper",
    data=excel_bytes,
    file_name="Pushpak_Kumar_Credit_Note_Mapping_Working_Paper.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

st.markdown("<div class='footer'>PUSHPAK KUMAR · GST AUTOMATION LAB · CREDIT NOTE MAPPING ENGINE</div>",unsafe_allow_html=True)
